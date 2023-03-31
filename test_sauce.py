from selenium import webdriver #web driver 
from webdriver_manager.chrome import ChromeDriverManager #web driver ı otomatik olarak bulur 
from time import sleep #kod u verilen zaman kadar bekletir
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait #bekleme işlemini süreli yerine (ikiside)
from selenium.webdriver.support import expected_conditions as ec #karakter bazlı olarak otomatik hale getirir
from selenium.webdriver.common.action_chains import ActionChains
import pytest #pytest paketi
from pathlib import Path
from datetime import date
import openpyxl
class Test_Sauce:
    def setup_method(self):
      self.web=webdriver.Chrome(ChromeDriverManager().install())
      self.web.maximize_window()
      self.web.get("https://www.saucedemo.com/")
      self.folderPath=str(date.today())
      Path(self.folderPath).mkdir(exist_ok=True)
    
    def teardown_method(self):
        self.web.quit()

    def getDate(self):
       exelFile=openpyxl.load_workbook("data/test_login.xlsx")
       selectedSheet=exelFile["Sayfa1"]
       maxRows=selectedSheet.max_row
       data=[]
       for i in range(2,maxRows+1):
          username=selectedSheet.cell(i,1).value
          password=selectedSheet.cell(i,2).value
          tupelData=(username,password)
          data.append(tupelData)
          return data
    
    def waitForElementVisible(self,locator,timeout=5):
       WebDriverWait(self.web,timeout).until(ec.visibility_of_all_elements_located(locator))
    
    @pytest.mark.parametrize("username","password",getDate())
    def test_saouce1(self,username,password):
       self.waitForElementVisible(By.ID,"user-name")
       usernameInput=self.web.find_element(By.ID,"user-name")
       self.waitForElementVisible(By.ID,"password")
       passwordInput=self.web.find_element(By.ID,"password")
       usernameInput.send_keys(username)
       passwordInput.send_keys(password)
       self.waitForElementVisible(By.ID,"login-button")
       loginbuttonInput=self.web.find_element(By.ID,"login-button")
       loginbuttonInput.click()
       errorMesage=self.web.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]")
       self.web.save_screenshot(f"{self.folderPath}/test_sauce_login-{username}-{password}.png")
       assert errorMesage.text=="Epic sadface: Username is required"

    
    def test_saouce2(self):
       self.waitForElementVisible(By.ID,"user-name")
       usernameInput=self.web.find_element(By.ID,"user-name")
       self.waitForElementVisible(By.ID,"password")
       passwordInput=self.web.find_element(By.ID,"password")
       usernameInput.send_keys("standard_user")
       passwordInput.send_keys("secret_sauce")
       self.waitForElementVisible(By.ID,"login-button")
       loginButtonInput=self.web.find_element(By.ID,"login-button")
       loginButtonInput.click()
       self.waitForElementVisible(By.XPATH,"/html/body/div/div/div/div[2]/div/div/div/div[1]/div[2]/div[2]/button")
       addcardInput=self.web.find_element(By.XPATH,"/html/body/div/div/div/div[2]/div/div/div/div[1]/div[2]/div[2]/button")
       addcardInput.click()
       self.waitForElementVisible(By.XPATH,"/html/body/div/div/div/div[2]/div/div/div/div[1]/div[2]/div[2]/button")
       removecardInput=self.web.find_element(By.XPATH,"/html/body/div/div/div/div[2]/div/div/div/div[1]/div[2]/div[2]/button")
       removecardInput.click()
       self.waitForElementVisible(By.XPATH,"/html/body/div/div/div/div[2]/div/div/div/div[1]/div[2]/div[2]/button")
       addcardInput2=self.web.find_element(By.XPATH,"/html/body/div/div/div/div[2]/div/div/div/div[1]/div[2]/div[2]/button")
       assert len(addcardInput2)>0
       



       
       
       
       